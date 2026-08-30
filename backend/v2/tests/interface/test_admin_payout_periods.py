"""Admin payout-period BFF routes."""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from backend.v2.contexts.finance.domain.payout_period import (
    PayoutPeriod,
    PayoutPeriodStateError,
    PayoutWarning,
    PersistedPayoutLine,
    PersistedUnpaidOccurrence,
)


def _dt(value: str) -> datetime:
    return datetime.fromisoformat(value).replace(tzinfo=UTC)


def _period(**overrides) -> PayoutPeriod:
    line = PersistedPayoutLine(
        occurrence_id="occ-pay-1",
        coach_id="coach-1",
        basis="actual",
        minutes=Decimal("90"),
        amount_minor=7500,
        currency="USD",
        rate_id="rate-1",
    )
    base = dict(
        period_id="pp-1",
        academy_id="acad",
        coach_id="coach-1",
        period_start=_dt("2026-05-01T00:00:00"),
        period_end=_dt("2026-06-01T00:00:00"),
        status="draft",
        currency="USD",
        total_minor=7500,
        lines=[line],
        unpaid_occurrence_ids=[],
        payout_warnings=[],
        generated_at=_dt("2026-06-01T10:00:00"),
        approved_at=None,
        paid_at=None,
    )
    base.update(overrides)
    return PayoutPeriod(**base)


def _warning(**overrides) -> PayoutWarning:
    base = dict(
        occurrence_id="occ-unpaid-1",
        reason="missing_session_price_for_percent_revenue",
        severity="blocking",
        message="Missing session price for percent-of-revenue pay.",
        occurred_at=_dt("2026-05-10T18:00:00"),
        session_id=None,
        session_title=None,
        coach_id="coach-1",
        repair_action="set_session_fee_and_recompute",
    )
    base.update(overrides)
    return PayoutWarning(**base)


class _GeneratePayoutPeriod:
    def __init__(self, period: PayoutPeriod) -> None:
        self.period = period
        self.calls: list[dict[str, object]] = []

    async def execute(self, **kwargs) -> PayoutPeriod:
        self.calls.append(kwargs)
        return self.period


class _ApprovePayoutPeriod:
    def __init__(self, period: PayoutPeriod) -> None:
        self.period = period
        self.calls: list[str] = []

    async def execute(self, *, period_id: str) -> PayoutPeriod:
        self.calls.append(period_id)
        if self.period.unpaid_occurrence_ids:
            raise PayoutPeriodStateError("unresolved unpaid occurrences remain")
        self.period = self.period.model_copy(
            update={"status": "approved", "approved_at": _dt("2026-06-02T10:00:00")}
        )
        return self.period


class _ApprovePayoutPeriodRaises:
    async def execute(self, *, period_id: str) -> PayoutPeriod:
        raise PayoutPeriodStateError("cannot approve payout period with unresolved payout warnings")


class _MarkPayoutPaid:
    def __init__(self, period: PayoutPeriod) -> None:
        self.period = period
        self.calls: list[object] = []

    async def execute(self, command) -> PayoutPeriod:
        self.calls.append(command)
        if self.period.status == "paid":
            return self.period
        self.period = self.period.model_copy(
            update={
                "status": "paid",
                "approved_at": self.period.approved_at or _dt("2026-06-02T10:00:00"),
                "paid_at": command.paid_at,
                "paid_method": command.method,
                "paid_amount_minor": command.amount_minor,
                "paid_reference": command.reference,
            }
        )
        return self.period


class _PayoutPeriodRepo:
    def __init__(self, period: PayoutPeriod | None) -> None:
        self.period = period

    async def find_by_id(self, period_id: str) -> PayoutPeriod | None:
        if self.period is None or self.period.period_id != period_id:
            return None
        return self.period


def _wire(admin_client, period: PayoutPeriod):
    admin_client.use_cases.generate_payout_period = _GeneratePayoutPeriod(period)
    admin_client.use_cases.approve_payout_period = _ApprovePayoutPeriod(period)
    admin_client.use_cases.mark_payout_paid = _MarkPayoutPaid(
        period.model_copy(update={"status": "approved", "approved_at": _dt("2026-06-02T10:00:00")})
    )
    admin_client.use_cases.payout_periods = _PayoutPeriodRepo(period)


def test_generate_payout_period_returns_review_payload(admin_client):
    period = _period()
    _wire(admin_client, period)

    response = admin_client.post(
        "/api/v2/admin/payout-periods/generate",
        json={
            "coach_id": "coach-1",
            "period_start": "2026-05-01T00:00:00Z",
            "period_end": "2026-06-01T00:00:00Z",
        },
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["period_id"] == "pp-1"
    assert body["status"] == "draft"
    assert body["total_amount_cents"] == 7500
    assert body["lines"][0]["occurrence_id"] == "occ-pay-1"
    assert body["lines"][0]["basis"] == "actual"
    assert admin_client.use_cases.generate_payout_period.calls[0]["academy_id"] == "acad"


def test_review_payout_period_detail(admin_client):
    _wire(admin_client, _period())

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1")

    assert response.status_code == 200, response.text
    assert response.json()["lines"][0]["amount_cents"] == 7500


def test_review_payout_period_detail_includes_warning_reason(admin_client):
    period = _period(
        unpaid_occurrence_ids=["occ-unpaid-1"],
        unpaid_occurrences=[
            PersistedUnpaidOccurrence(
                occurrence_id="occ-gap",
                reason="rate_gap",
                detail="Coach pay-rate history has a gap.",
                unresolved=True,
            )
        ],
        payout_warnings=[_warning()],
    )
    _wire(admin_client, period)
    admin_client.use_cases.describe_payout_occurrences = AsyncDescribe(
        {
            "occ-unpaid-1": {
                "occurred_at": _dt("2026-05-10T18:00:00"),
                "session_id": "sess-1",
                "session_title": "Junior Squad",
            }
        }
    )

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["unpaid_occurrence_ids"] == ["occ-unpaid-1"]
    assert body["payout_warnings"][0]["reason"] == "missing_session_price_for_percent_revenue"
    assert body["payout_warnings"][0]["session_title"] == "Junior Squad"
    by_occurrence = {row["occurrence_id"]: row for row in body["unpaid_occurrences"]}
    assert by_occurrence["occ-unpaid-1"]["reason"] == "missing_session_price_for_percent_revenue"
    assert by_occurrence["occ-unpaid-1"]["repair_action"] == "set_session_fee_and_recompute"
    assert by_occurrence["occ-gap"]["reason"] == "rate_gap"
    assert by_occurrence["occ-gap"]["detail"] == "Coach pay-rate history has a gap."
    assert by_occurrence["occ-gap"]["unresolved"] is True


class AsyncDescribe:
    def __init__(self, rows: dict[str, dict[str, object]]) -> None:
        self.rows = rows

    async def __call__(self, occurrence_ids: list[str]) -> dict[str, dict[str, object]]:
        return {occ_id: self.rows[occ_id] for occ_id in occurrence_ids if occ_id in self.rows}


def test_approve_and_mark_paid_with_metadata_are_idempotent(admin_client):
    _wire(admin_client, _period())

    approved = admin_client.post("/api/v2/admin/payout-periods/pp-1/approve")
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "approved"

    paid = admin_client.post(
        "/api/v2/admin/payout-periods/pp-1/mark-paid",
        json={
            "method": "bank_transfer",
            "paid_at": "2026-06-03T10:00:00Z",
            "amount_cents": 7500,
            "reference": "ach-456",
        },
    )
    assert paid.status_code == 200, paid.text
    body = paid.json()
    assert body["status"] == "paid"
    assert body["paid_method"] == "bank_transfer"
    assert body["paid_amount_cents"] == 7500
    assert body["paid_reference"] == "ach-456"

    again = admin_client.post(
        "/api/v2/admin/payout-periods/pp-1/mark-paid",
        json={
            "method": "check",
            "paid_at": "2026-06-04T10:00:00Z",
            "amount_cents": 9999,
            "reference": "retry",
        },
    )
    assert again.status_code == 200, again.text
    assert again.json()["paid_method"] == "bank_transfer"
    assert again.json()["paid_reference"] == "ach-456"


def test_approve_payout_period_blocks_unresolved_unpaid_occurrences(admin_client):
    period = _period(
        unpaid_occurrence_ids=["occ-gap"],
        unpaid_occurrences=[
            PersistedUnpaidOccurrence(
                occurrence_id="occ-gap",
                reason="rate_gap",
                detail="Coach pay-rate history has a gap.",
                unresolved=True,
            )
        ],
        payout_warnings=[_warning(occurrence_id="occ-warning")],
    )
    _wire(admin_client, period)
    admin_client.use_cases.approve_payout_period = _ApprovePayoutPeriodRaises()

    response = admin_client.post("/api/v2/admin/payout-periods/pp-1/approve")

    assert response.status_code == 400
    assert "unresolved payout warnings" in response.text


def test_printable_payslip_contains_period_and_lines(admin_client):
    period = _period(
        status="paid",
        approved_at=_dt("2026-06-02T10:00:00"),
        paid_at=_dt("2026-06-03T10:00:00"),
        paid_method="bank_transfer",
        paid_amount_minor=7500,
        paid_reference="ach-456",
    )
    _wire(admin_client, period)

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1/payslip")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["printable"] is True
    assert body["period"]["period_id"] == "pp-1"
    assert body["period"]["paid_reference"] == "ach-456"
    assert body["lines"][0]["occurrence_id"] == "occ-pay-1"


def test_payout_period_export_includes_warning_details(admin_client):
    period = _period(
        unpaid_occurrence_ids=["occ-unpaid-1"],
        payout_warnings=[_warning()],
    )
    _wire(admin_client, period)
    admin_client.use_cases.describe_payout_occurrences = AsyncDescribe(
        {
            "occ-unpaid-1": {
                "occurred_at": _dt("2026-05-10T18:00:00"),
                "session_id": "sess-1",
                "session_title": "Junior Squad",
            }
        }
    )

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1/export")

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any("missing_session_price_for_percent_revenue" in row for row in rows)
    assert any("set_session_fee_and_recompute" in row for row in rows)


def test_review_payout_period_surfaces_replaced_occurrence(admin_client):
    """A displaced scheduled coach's period must explain the exclusion (#228)."""
    period = _period(
        unpaid_occurrences=[
            PersistedUnpaidOccurrence(
                occurrence_id="occ-replaced",
                reason="replaced_by_actual_coach",
                detail="Scheduled coach was replaced; this occurrence was attributed to coach coach-2.",
                unresolved=False,
                attributed_coach_id="coach-2",
            )
        ],
    )
    _wire(admin_client, period)
    admin_client.use_cases.describe_payout_occurrences = AsyncDescribe(
        {
            "occ-replaced": {
                "occurred_at": _dt("2026-05-10T18:00:00"),
                "session_id": "sess-1",
                "session_title": "Junior Squad",
            }
        }
    )

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1")

    assert response.status_code == 200, response.text
    body = response.json()
    row = next(r for r in body["unpaid_occurrences"] if r["occurrence_id"] == "occ-replaced")
    assert row["reason"] == "replaced_by_actual_coach"
    assert row["attributed_coach_id"] == "coach-2"
    assert row["unresolved"] is False
    # Enrichment must reach structured rows, not just ``unpaid_occurrence_ids``.
    assert row["session_title"] == "Junior Squad"
    # A replacement is not a repair item, so it stays out of the blocking list.
    assert body["unpaid_occurrence_ids"] == []
    assert body["total_amount_cents"] == 7500


def test_payout_period_export_includes_replaced_occurrence(admin_client):
    period = _period(
        unpaid_occurrences=[
            PersistedUnpaidOccurrence(
                occurrence_id="occ-replaced",
                reason="replaced_by_actual_coach",
                detail="Scheduled coach was replaced.",
                unresolved=False,
                attributed_coach_id="coach-2",
            )
        ],
    )
    _wire(admin_client, period)

    response = admin_client.get("/api/v2/admin/payout-periods/pp-1/export")

    assert response.status_code == 200, response.text
    rows = list(load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))
    assert any("replaced_by_actual_coach" in row for row in rows)
    assert any("attributed to coach coach-2" in row for row in rows)


def test_payout_period_routes_wrong_persona_404(coach_on_admin_client):
    response = coach_on_admin_client.get("/api/v2/admin/payout-periods/pp-1")
    assert response.status_code == 404
