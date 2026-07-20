"""One-shot importer for BLno Badminton Training spreadsheet.

Reads /tmp/blno.xlsx (or provided path) and seeds the DB.

WARNING: drops existing transactional collections (sessions, students,
enrollments, payments, expenses, attendance, lesson_plans, progress_notes,
coach_payouts, payout_rules, move_log, messages, notifications, invites).
Keeps only the configured admin account, then adds the two real coaches
and all parents from the sheet.
"""
import os
import re
import sys
import asyncio
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(ROOT / ".env")

import bcrypt  # noqa: E402
import openpyxl  # noqa: E402
from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402


SOURCE = os.environ.get("BLNO_XLSX", "/tmp/blno.xlsx")
COACH_PASSWORD = os.environ.get("BLNO_COACH_PASSWORD")
PARENT_PASSWORD = os.environ.get("BLNO_PARENT_PASSWORD")
if not COACH_PASSWORD or not PARENT_PASSWORD:
    sys.exit(
        "BLNO_COACH_PASSWORD and BLNO_PARENT_PASSWORD must be set "
        "before running import_blno.py"
    )


def hp(p: str) -> str:
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def period_from_date(dt) -> str:
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m")
    raw = str(dt or "").strip()
    for fmt in ("%b-%Y", "%B-%Y", "%Y-%m"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m")
        except ValueError:
            pass
    return raw


SESSION_RE = re.compile(r"(?P<day>\w+)\s+(?P<start>\d{1,2}:\d{2}\s*[AP]M)\s*[–-]\s*(?P<end>\d{1,2}:\d{2}\s*[AP]M)\s+(?P<level>\w+)\(Coach\s*-\s*(?P<coach>\w+)\)")


def parse_session_name(name: str):
    m = SESSION_RE.match((name or "").strip())
    if not m:
        return None
    d = m.groupdict()
    return {
        "day": d["day"],
        "start_time": _to24(d["start"]),
        "end_time": _to24(d["end"]),
        "skill_level": d["level"].lower(),
        "coach_name": d["coach"],
        "raw_name": name.strip(),
    }


def _to24(s: str) -> str:
    return datetime.strptime(s.strip(), "%I:%M %p").strftime("%H:%M")


def split_name(full: str):
    parts = (full or "").strip().split()
    if len(parts) <= 1:
        return parts[0] if parts else "", ""
    return parts[0], " ".join(parts[1:])


def fmt_phone(p) -> str:
    if p is None:
        return ""
    s = str(p).replace(".0", "").replace("+", "").replace("-", "").replace(" ", "").replace("(", "").replace(")", "")
    return s


async def main():
    mongo_url = os.environ["MONGO_URL"]
    db = AsyncIOMotorClient(mongo_url)[os.environ["DB_NAME"]]

    print("Loading spreadsheet:", SOURCE)
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    def rows(sheet, header_row=0):
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        if header_row >= len(data):
            return []
        headers = data[header_row]
        out = []
        for r in data[header_row + 1:]:
            if all(v is None for v in r):
                continue
            out.append({h: v for h, v in zip(headers, r) if h})
        return out

    # 1) Drop transactional data; keep only the configured admin user.
    for c in ("sessions", "students", "enrollments", "payments", "expenses",
              "attendance", "lesson_plans", "progress_notes", "coach_payouts",
              "payout_rules", "move_log", "messages", "notifications", "invites",
              "audit_logs"):
        await db[c].drop()
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@badminton.app").lower()
    await db.users.delete_many({"email": {"$ne": admin_email}})
    print("Cleared transactional collections.")

    # 2) Inputs sheet — pricing, capacity, payout %, expenses
    inputs_ws = wb["Inputs"]
    inputs_kv = {}
    for r in inputs_ws.iter_rows(values_only=True):
        if r and r[0] and isinstance(r[0], str):
            inputs_kv[r[0].strip()] = r[1:]
    capacity_per_session = 15
    beginner_price = 60
    intermediate_price = 70
    # Try to extract numbers from the inputs sheet
    for k, vs in inputs_kv.items():
        kl = k.lower()
        v = vs[0] if vs else None
        if isinstance(v, (int, float)):
            if "capacity per session" in kl:
                capacity_per_session = int(v)
            elif "beginner price" in kl:
                beginner_price = int(v)
            elif "intermediate price" in kl:
                intermediate_price = int(v)
    print(f"Pricing: Beginner=${beginner_price}, Intermediate=${intermediate_price}, Cap={capacity_per_session}")

    # Months / Rent / Other Exp rows — search "Months" header inside Inputs
    months_data = []  # list of (period, rent, other)
    inp_rows = list(inputs_ws.iter_rows(values_only=True))
    for i, r in enumerate(inp_rows):
        if r and isinstance(r[0], str) and r[0].strip().lower() == "month":
            for r2 in inp_rows[i + 1:]:
                if r2 and r2[0]:
                    months_data.append({
                        "period": period_from_date(r2[0]),
                        "rent": float(r2[1]) if r2[1] else 0,
                        "other": float(r2[2]) if r2[2] else 0,
                    })
            break

    # 3) Coaches — seed Gowtham & Kishore
    coach_emails = {"Gowtham": "gowtham@blno.academy", "Kishore": "kishore@blno.academy"}
    coach_ids = {}
    for cname, cemail in coach_emails.items():
        existing = await db.users.find_one({"email": cemail})
        if existing:
            coach_ids[cname] = str(existing["_id"])
        else:
            r = await db.users.insert_one({
                "email": cemail, "password_hash": hp(COACH_PASSWORD),
                "name": cname, "phone": "", "role": "coach", "status": "active",
                "created_at": now(), "updated_at": now(),
            })
            coach_ids[cname] = str(r.inserted_id)
    print("Coaches:", coach_ids)

    # 4) Sessions — distinct session names from Roster
    roster_rows = rows("Roster")
    distinct_sessions = {}
    for r in roster_rows:
        name = (r.get("Session") or "").strip()
        if not name:
            continue
        if name not in distinct_sessions:
            p = parse_session_name(name)
            if not p:
                continue
            price = intermediate_price if p["skill_level"] == "intermediate" else beginner_price
            distinct_sessions[name] = {
                "name": name, "skill_level": p["skill_level"], "age_group": "All",
                "start_date": "2026-04-01", "end_date": "2026-12-31",
                "days_of_week": [p["day"][:3]], "start_time": p["start_time"],
                "end_time": p["end_time"], "location": "Court", "max_students": capacity_per_session,
                "monthly_price": float(price), "coach_id": coach_ids.get(p["coach_name"]),
                "status": "active", "is_deleted": False, "created_at": now(),
            }
    session_ids = {}
    for name, doc in distinct_sessions.items():
        r = await db.sessions.insert_one(doc)
        session_ids[name] = str(r.inserted_id)
    print(f"Sessions: {len(session_ids)} created")

    # 5) Payout rules — 30% revenue for each coach
    for cname, cid in coach_ids.items():
        await db.payout_rules.insert_one({
            "coach_id": cid, "rule_type": "revenue_percentage",
            "value": 30, "is_active": True, "created_at": now(),
        })

    # 6) Form_Responses — index by Child Full Name for medical/waiver/t-shirt/exp
    form_data = {}
    for r in rows("Form_Responses"):
        name = (r.get("Child Full Name") or "").strip()
        if not name:
            continue
        form_data[name.lower()] = r

    # 7) Parents + Students + Enrollments
    parent_id_by_email = {}
    student_ids_by_name = {}
    enrollment_ids = {}  # (student_id, session_id) → enrollment_id
    for r in roster_rows:
        child = (r.get("Child Name") or "").strip()
        parent_name = (r.get("Parent") or "").strip()
        phone = fmt_phone(r.get("Phone"))
        email = (r.get("Email") or "").strip().lower()
        skill = (r.get("Skill") or "Beginner").strip().lower()
        session_name = (r.get("Session") or "").strip()
        status = (r.get("Status") or "Active").strip().lower()
        billing_type_raw = (r.get("Billing Type") or "Standard").strip()
        if not child or not email:
            continue
        # Parent user
        if email not in parent_id_by_email:
            existing = await db.users.find_one({"email": email})
            if existing:
                parent_id_by_email[email] = str(existing["_id"])
            else:
                rr = await db.users.insert_one({
                    "email": email, "password_hash": hp(PARENT_PASSWORD),
                    "name": parent_name or email.split("@")[0],
                    "phone": phone, "role": "parent", "status": "active",
                    "created_at": now(), "updated_at": now(),
                })
                parent_id_by_email[email] = str(rr.inserted_id)
        parent_id = parent_id_by_email[email]

        # Student
        first, last = split_name(child)
        form = form_data.get(child.lower(), {})
        def _s(v):
            return ("" if v is None else str(v)).strip()
        emergency_raw = _s(form.get("Emergency Contact Name and Phone Number"))
        medical = _s(form.get("Any medical condition, allergy, or injury we should know about?"))
        tshirt = _s(form.get("T-shirt Size"))
        prev_exp = _s(form.get("Previous Badminton Experience"))
        waiver = bool(form.get("Waiver Agreement"))
        age = int(form.get("Child Age") or 0)
        dob = ""
        if age:
            try:
                dob = (datetime.now() - timedelta(days=365 * age)).strftime("%Y-%m-%d")
            except Exception:
                dob = ""
        ec_name = emergency_raw.split(":")[0].split("-")[0].strip()[:80] if emergency_raw else ""
        ec_phone = re.findall(r"\d[\d\s\-]{8,}", emergency_raw)
        ec_phone = ec_phone[0] if ec_phone else ""
        stu_doc = {
            "first_name": first, "last_name": last, "dob": dob, "age": age,
            "skill_level": skill, "emergency_contact_name": ec_name,
            "emergency_contact_phone": ec_phone, "medical_notes": medical,
            "waiver_accepted": waiver,
            "waiver_date": now() if waiver else None,
            "t_shirt_size": tshirt, "previous_experience": prev_exp,
            "parent_user_id": parent_id,
            "status": "active" if status == "active" else status,
            "is_deleted": False, "created_at": now(),
        }
        sr = await db.students.insert_one(stu_doc)
        student_id = str(sr.inserted_id)
        student_ids_by_name[child.lower()] = student_id

        # Enrollment
        sid = session_ids.get(session_name)
        if not sid:
            continue
        # Normalize billing_type
        bt = "Standard"
        b_low = billing_type_raw.lower()
        if "no" in b_low and "charge" in b_low:
            bt = "NoCharge"
        elif "waiv" in b_low:
            bt = "Waived"
        en_doc = {
            "session_id": sid, "student_id": student_id, "parent_user_id": parent_id,
            "billing_type": bt, "approval_status": "approved",
            "status": "active", "enrolled_at": now(), "is_deleted": False,
        }
        er = await db.enrollments.insert_one(en_doc)
        enrollment_ids[(student_id, sid)] = str(er.inserted_id)

        # Payments for Apr/May from roster
        for col_pay, col_due, period in (
            ("Apr-2026 Pay", "Apr-2026 Due", "2026-04"),
            ("May-2026 Pay", "May-2026 Due", "2026-05"),
        ):
            paid_amt = r.get(col_pay)
            due_amt = r.get(col_due)
            if paid_amt is None and due_amt is None:
                continue
            amount = float(distinct_sessions[session_name]["monthly_price"])
            if bt != "Standard":
                continue  # No payment row for no-charge
            paid_v = float(paid_amt) if paid_amt else 0
            due_v = float(due_amt) if due_amt else 0
            status_p = "paid" if paid_v >= amount and due_v <= 0 else "pending"
            doc = {
                "parent_user_id": parent_id, "student_id": student_id,
                "enrollment_id": str(er.inserted_id), "session_id": sid,
                "period": period, "amount": amount,
                "discount": 0, "final_amount": amount - due_v if status_p == "paid" else amount,
                "status": status_p,
                "payment_date": now() if status_p == "paid" else None,
                "payment_method": "Zelle" if status_p == "paid" else None,
                "marked_by": None, "notes": "imported", "is_deleted": False,
                "created_at": now(),
            }
            # If partial paid (paid_v between 0 and amount)
            if 0 < paid_v < amount:
                doc["final_amount"] = amount
                doc["status"] = "pending"
            await db.payments.insert_one(doc)

    print(f"Parents: {len(parent_id_by_email)}, Students: {len(student_ids_by_name)}, "
          f"Enrollments: {len(enrollment_ids)}")

    # 8) Expenses (Rent + Other for each month)
    for m in months_data:
        if m["rent"]:
            await db.expenses.insert_one({
                "category": "Court rental", "description": "Monthly rent",
                "amount": m["rent"], "date": f"{m['period']}-01",
                "paid_to": "Landlord", "status": "paid", "notes": "imported",
                "is_deleted": False, "created_by": None, "created_at": now(),
            })
        if m["other"]:
            await db.expenses.insert_one({
                "category": "Miscellaneous", "description": "Other expenses",
                "amount": m["other"], "date": f"{m['period']}-01",
                "paid_to": "", "status": "paid", "notes": "imported",
                "is_deleted": False, "created_by": None, "created_at": now(),
            })

    # 9) Attendance from Audit_Log (mark_attendance entries)
    audit_added = 0
    for r in rows("Audit_Log"):
        if r.get("Action") != "mark_attendance":
            continue
        kid = (r.get("Target") or "").strip()
        stu_id = student_ids_by_name.get(kid.lower())
        if not stu_id:
            continue
        scope = r.get("Scope")
        meta = r.get("Meta") or "{}"
        try:
            meta_d = json.loads(meta) if isinstance(meta, str) else {}
        except Exception:
            meta_d = {}
        session_name = meta_d.get("session", "")
        sid = session_ids.get(session_name)
        if not sid:
            continue
        date_v = scope if isinstance(scope, datetime) else None
        date_s = date_v.strftime("%Y-%m-%d") if date_v else now()[:10]
        status_v = (r.get("After") or "Present").strip().lower()
        # Normalize make-up / makeup
        if "make" in status_v:
            status_v = "make_up"
        await db.attendance.update_one(
            {"session_id": sid, "student_id": stu_id, "date": date_s},
            {"$set": {
                "session_id": sid, "student_id": stu_id, "date": date_s,
                "status": status_v, "notes": meta_d.get("note", ""),
                "marked_by": coach_ids.get("Gowtham"),  # default
                "marked_at": now(),
            }}, upsert=True)
        audit_added += 1
    print(f"Attendance entries imported: {audit_added}")

    # 10) Move_Log → move_log + apply override / permanent
    for r in rows("Move_Log"):
        kid = (r.get("Kid") or "").strip()
        if not kid:
            continue
        stu_id = student_ids_by_name.get(kid.lower())
        if not stu_id:
            continue
        from_name = (r.get("From") or "").strip()
        to_name = (r.get("To") or "").strip()
        eff = r.get("Effective Month")
        if isinstance(eff, datetime):
            eff_period = eff.strftime("%Y-%m")
        else:
            eff_period = str(eff or "")
        from_sid = session_ids.get(from_name)
        to_sid = session_ids.get(to_name)
        if not (from_sid and to_sid):
            continue
        await db.move_log.insert_one({
            "student_id": stu_id, "from_session_id": from_sid, "to_session_id": to_sid,
            "effective_month": eff_period, "permanent": True,
            "note": "imported", "moved_by": None, "moved_at": now(),
        })

    print("✓ Import complete.")

if __name__ == "__main__":
    asyncio.run(main())
